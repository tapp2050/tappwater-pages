from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ── Color constants ──
DARK_BLUE = PatternFill('solid', fgColor='1F3864')
LIGHT_GREY = PatternFill('solid', fgColor='F2F2F2')
WHITE = PatternFill('solid', fgColor='FFFFFF')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
RED_FILL = PatternFill('solid', fgColor='FCE4EC')
YELLOW_FILL = PatternFill('solid', fgColor='FFF9C4')
BLUE_FILL = PatternFill('solid', fgColor='DCEEFB')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
CURRENCY_FMT = '$#,##0.00'
NUMBER_FMT = '#,##0'
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9')
)

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = DARK_BLUE
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'

def style_body(ws, num_rows, num_cols, start_row=2):
    for row in range(start_row, num_rows + 1):
        fill = LIGHT_GREY if (row - start_row) % 2 == 0 else WHITE
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if cell.font == Font():
                cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

def auto_width(ws, num_cols, max_width=45):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, max_width)

# ════════════════════════════════════════════
# SHEET 1: Content Plan
# ════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Content Plan"

headers1 = ["Phase", "Priority", "Page Type", "Page Title", "URL Slug", "Primary Keyword",
            "Primary KW Vol", "Secondary Keywords", "Combined Cluster Vol",
            "Content Angle", "Product Tie-in", "Est. Clicks/mo (Top 3)", "Status"]
ws1.append(headers1)

content_pages = [
    # PHASE 1
    ["1", "P1", "Collection", "Shower Filters Australia — Filtered Shower Heads for Hair & Skin",
     "/collections/shower-filter", "shower filter", 12100,
     "shower filter australia, shower water filter, filtered shower head, shower head filter, shower head water filter, shower water filter australia",
     44000, "Main shower filter collection page. Position around hair + skin benefits. Target all head terms in this wide-open category.",
     "ShowerPro", 2500, "Not Started"],

    ["1", "P1", "Collection", "Countertop Reverse Osmosis Water Filter — No Installation Required",
     "/collections/countertop-reverse-osmosis", "countertop reverse osmosis water filter", 210,
     "benchtop reverse osmosis water filter, benchtop reverse osmosis water filter australia, countertop water purifier, portable reverse osmosis, countertop ro system",
     670, "Launch page for new RO product. Lead with plug-and-play, no plumber needed. Also target broader RO cluster as the countertop alternative.",
     "Countertop RO", 400, "Not Started"],

    ["1", "P1", "Product", "EcoPro — Best Tap Water Filter Australia",
     "/products/tap-water-filter", "tap water filter", 12100,
     "tap filter, water filter tap, faucet water filter, best tap filter, on tap water filter, screw on tap filter, water filter for tap australia",
     29180, "Optimise existing product page. Strengthen 'best tap filter' #1 position, push 'tap water filter' from #9 to top 5.",
     "EcoPro", 1500, "Not Started"],

    ["1", "P1", "Landing Page", "Best Water Filter Australia 2026",
     "/blogs/best-water-filter-australia", "best water filter australia", 2400,
     "water filter australia, best tap water filter australia",
     6210, "Comparison/roundup page featuring all Tapp products plus competitors for credibility. SEO listicle format.",
     "All", 600, "Not Started"],

    ["1", "P1", "Landing Page", "Best Shower Filter Australia 2026",
     "/blogs/best-shower-filter-australia", "best shower filter australia", 480,
     "best shower filter, shower filter australia",
     2170, "Comparison page for shower filters. Include ShowerPro, Well Verti, Filtered Beauty, Fettle Effect. Honest reviews.",
     "ShowerPro", 300, "Not Started"],

    ["1", "P1", "Landing Page", "Best Reverse Osmosis Water Filter Australia 2026",
     "/blogs/best-reverse-osmosis-water-filter-australia", "best reverse osmosis water filter australia", 320,
     "reverse osmosis water filter australia, reverse osmosis water filter, reverse osmosis water filter system",
     9720, "Comparison page. Position Tapp countertop RO as the no-plumber option. Compare to Waterdrop, under-sink systems.",
     "Countertop RO", 800, "Not Started"],

    # PHASE 2
    ["2", "P1", "Guide", "Tap Water in Sydney: What's In It & Should You Filter It?",
     "/blogs/tap-water-sydney", "tap water sydney", 390,
     "sydney water quality, can you drink tap water in sydney, is sydney water safe to drink, hard water sydney",
     1660, "Comprehensive guide to Sydney water. Cover source, treatment, contaminants, hardness. Link to all products. Spike-ready for water incidents.",
     "All", 200, "Not Started"],

    ["2", "P1", "Guide", "Tap Water in Melbourne: What's In It & Should You Filter It?",
     "/blogs/tap-water-melbourne", "tap water melbourne", 590,
     "melbourne water quality, can you drink tap water in melbourne, is melbourne water safe to drink, hard water melbourne",
     1740, "Melbourne-specific water guide. Spiked to 1,900 during Nov '25 water incident. Same template as Sydney.",
     "All", 220, "Not Started"],

    ["2", "P1", "Guide", "Tap Water in Brisbane: What's In It & Should You Filter It?",
     "/blogs/tap-water-brisbane", "tap water brisbane", 260,
     "brisbane water quality, can you drink tap water in brisbane, is tap water safe to drink in brisbane, is brisbane water safe to drink",
     890, "Brisbane-specific. Has spiked to 1,000+ during water quality incidents.",
     "All", 120, "Not Started"],

    ["2", "P2", "Guide", "Tap Water in Perth: Hard Water Capital of Australia",
     "/blogs/tap-water-perth", "hard water perth", 210,
     "tap water perth, perth water quality, is perth water safe to drink, reverse osmosis water filter perth",
     720, "Perth = hard water angle. Direct link to ShowerPro for shower hardness + RO for drinking water.",
     "ShowerPro + RO", 90, "Not Started"],

    ["2", "P2", "Guide", "Can You Drink Tap Water in Australia?",
     "/blogs/can-you-drink-tap-water-in-australia", "can you drink tap water in australia", 590,
     "is tap water safe to drink, is tap water safe in australia, tap water contaminants",
     880, "National overview / hub page linking to all city pages. Covers what AU water treatment does and doesn't remove.",
     "All", 120, "Not Started"],

    ["2", "P2", "Guide", "Tap Water in Canberra: Water Quality & What to Know",
     "/blogs/tap-water-canberra", "canberra water quality", 110,
     "", 110, "Canberra has $2.88 CPC — highest in entire city cluster. Small but high commercial intent.",
     "All", 15, "Not Started"],

    ["2", "P1", "Guide", "PFAS & Forever Chemicals in Australian Water",
     "/blogs/pfas-water-australia", "pfas water filter", 480,
     "pfas in water, pfas australia, pfas filter, forever chemicals australia, pfas in drinking water, remove pfas from water",
     1640, "Spike-ready content. Explain PFAS, where found in AU, which filters remove them. Key message: RO removes PFAS, most tap filters don't.",
     "Countertop RO", 150, "Not Started"],

    ["2", "P1", "Guide", "Fluoride in Australian Tap Water: Facts, Risks & How to Remove It",
     "/blogs/fluoride-in-water-australia", "fluoride in water", 2400,
     "fluoride water filter, water filter that removes fluoride, fluoride filter, fluoride in water australia, remove fluoride from water",
     4560, "High-volume, controversial. Present facts neutrally. Key message: only RO removes fluoride. Links to RO product.",
     "Countertop RO", 400, "Not Started"],

    ["2", "P2", "Guide", "Microplastics in Drinking Water: Should You Be Worried?",
     "/blogs/microplastics-water-filter", "microplastics water filter", 320,
     "microplastics filter, microplastics in water, microplastics in drinking water, microplastics in tap water, microplastics in bottled water, remove microplastics from water",
     730, "Fastest growing cluster (doubling in 2 months). Cover what they are, recent studies, which filters remove them. First-mover advantage.",
     "All", 100, "Not Started"],

    ["2", "P2", "Guide", "Hard Water in Australia: Where It Is, What It Does & How to Fix It",
     "/blogs/hard-water-australia", "hard water", 1900,
     "hard water australia, hard water filter, water hardness australia, limescale, water softener, hard water hair, hard water eczema, hard water perth, hard water melbourne, hard water sydney",
     8414, "Comprehensive hard water guide with AU map by city. Link to ShowerPro for shower, RO for drinking. Hub for city pages.",
     "ShowerPro + RO", 600, "Not Started"],

    ["2", "P2", "Guide", "Chlorine in Tap Water: Why It's There & Should You Filter It?",
     "/blogs/chlorine-in-tap-water", "chlorine in water", 260,
     "chlorine in tap water, chlorine water filter, remove chlorine from water, chlorine skin, chlorine hair damage",
     620, "Practical guide. Every Tapp product removes chlorine — cross-sell opportunity for all three products.",
     "All", 80, "Not Started"],

    # PHASE 3
    ["3", "P2", "Guide", "Shower Filter for Hard Water: Does It Actually Work?",
     "/blogs/shower-filter-hard-water", "shower filter for hard water", 1000,
     "hard water shower",
     1040, "Problem-solution content. What shower filters can/can't do for hard water. Link to ShowerPro.",
     "ShowerPro", 150, "Not Started"],

    ["3", "P2", "Guide", "Shower Filter for Chlorine: Protect Your Hair & Skin",
     "/blogs/shower-filter-chlorine", "shower filter for chlorine", 390,
     "chlorine shower filter, chlorine hair damage, chlorine skin",
     920, "Health/beauty angle. Chlorine damage to hair and skin, how shower filters help. Before/after messaging.",
     "ShowerPro", 120, "Not Started"],

    ["3", "P2", "Guide", "Can a Shower Filter Help Eczema?",
     "/blogs/shower-filter-eczema", "shower filter for eczema", 70,
     "hard water eczema, water filter for eczema",
     120, "Small volume but SPIKING (40→140 Jan '26). High emotional intent. Cite studies and customer reviews.",
     "ShowerPro", 30, "Not Started"],

    ["3", "P3", "Guide", "Shower Filter for Hair Loss: What the Science Says",
     "/blogs/shower-filter-hair-loss", "shower filter for hair loss", 30,
     "hard water hair loss, hard water hair",
     190, "Spiking from near-zero. Honest, science-backed content. Growing awareness cluster.",
     "ShowerPro", 25, "Not Started"],

    ["3", "P2", "Comparison", "Reverse Osmosis vs Tap Water Filter: Which Do You Need?",
     "/blogs/reverse-osmosis-vs-tap-filter", "reverse osmosis vs water filter", 20,
     "water filter comparison, tap water vs filtered water",
     160, "Decision-stage content. When do you need RO vs a simple tap filter? Guide people to the right Tapp product.",
     "All", 25, "Not Started"],

    ["3", "P2", "Comparison", "Countertop vs Under-Sink Reverse Osmosis: Pros, Cons & Which to Buy",
     "/blogs/countertop-vs-under-sink-ro", "countertop reverse osmosis", 210,
     "under sink reverse osmosis",
     470, "Direct comparison. Tapp's no-install advantage vs under-sink hassle. Perfect for renters and apartments.",
     "Countertop RO", 60, "Not Started"],

    ["3", "P3", "Guide", "Best Water Filter for Renters (No Installation Required)",
     "/blogs/water-filter-for-renters", "water filter for renters", 10,
     "reverse osmosis for renters, water filter no installation",
     50, "Tiny volume now but perfectly aligned with Tapp USP. Own before it grows. Targets emerging 'RO for renters' related searches.",
     "All", 10, "Not Started"],

    ["3", "P2", "Guide", "Water Filter for Caravans: What to Look For",
     "/blogs/water-filter-caravan", "water filter caravan", 2900,
     "water filter camping",
     3780, "Huge AU-specific volume. Even without a dedicated caravan product, capture traffic and position countertop RO as an option.",
     "Countertop RO", 300, "Not Started"],

    ["3", "P3", "Comparison", "Tapp Water vs Brita: Which Filter Is Better?",
     "/blogs/tapp-vs-brita", "brita water filter", 18100,
     "brita alternative, brita filter review, brita replacement filter",
     18220, "Competitor comparison. Honest head-to-head. Capture people considering or already using Brita.",
     "EcoPro", 200, "Not Started"],

    ["3", "P3", "Guide", "Water Filters at Bunnings: What They Stock & What They Miss",
     "/blogs/water-filter-bunnings", "water filter bunnings", 12100,
     "reverse osmosis water filter bunnings",
     12690, "Intercept Bunnings shoppers. Cover what Bunnings stocks vs what they don't (countertop RO, subscription filters).",
     "All", 400, "Not Started"],

    ["3", "P3", "Guide", "Is Bottled Water Better Than Tap? The Truth for Australians",
     "/blogs/bottled-water-vs-tap-water", "is bottled water better than tap", 90,
     "bottled water plastic, microplastics in bottled water, bottled water vs filtered water, bottled water australia, water filter vs bottled water",
     530, "Brand story content. Cost comparison, environmental angle, microplastics in bottles. Social sharing + link building.",
     "All", 50, "Not Started"],

    ["3", "P3", "Guide", "Water Filter for Your Office: Best Options 2026",
     "/blogs/water-filter-office", "water filter office", 210,
     "", 210, "Highest CPC use-case ($1.86). Position countertop RO as the no-install office solution.",
     "Countertop RO", 30, "Not Started"],

    ["3", "P3", "Guide", "Lead in Australian Tap Water: Should You Be Concerned?",
     "/blogs/lead-in-water-australia", "lead in water", 90,
     "lead water filter",
     160, "Lead water filter term TRIPLING (30→140). Another RO advantage piece.",
     "Countertop RO", 20, "Not Started"],

    ["3", "P3", "Guide", "Vitamin C Shower Filter: Do They Actually Work?",
     "/blogs/vitamin-c-shower-filter", "vitamin c shower filter", 210,
     "", 210, "Informational content comparing vitamin C media to KDF/calcium sulfite (what ShowerPro uses).",
     "ShowerPro", 30, "Not Started"],
]

for row in content_pages:
    ws1.append(row)

style_header(ws1, len(headers1))
style_body(ws1, len(content_pages) + 1, len(headers1))

# Format number columns
for row in range(2, len(content_pages) + 2):
    ws1.cell(row=row, column=7).number_format = NUMBER_FMT
    ws1.cell(row=row, column=9).number_format = NUMBER_FMT
    ws1.cell(row=row, column=12).number_format = NUMBER_FMT

# Bold phase separators
current_phase = None
for row in range(2, len(content_pages) + 2):
    phase = ws1.cell(row=row, column=1).value
    if phase != current_phase:
        current_phase = phase
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row, column=col).font = BOLD_FONT

# Column widths
widths1 = [8, 8, 12, 55, 45, 35, 12, 60, 15, 60, 15, 15, 12]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════
# SHEET 2: Keyword Database
# ════════════════════════════════════════════
ws2 = wb.create_sheet("Keyword Database")

headers2 = ["Keyword", "Monthly Volume (AU)", "CPC (USD)", "Competition", "Cluster",
            "Trend Direction", "Current Tapp Ranking", "Target Page"]
ws2.append(headers2)

trend_fills = {
    "Rising": GREEN_FILL,
    "Declining": RED_FILL,
    "Volatile": YELLOW_FILL,
    "Spiking": BLUE_FILL,
    "Stable": WHITE,
}

keywords = [
    # Tap Filter
    ["water filter", 33100, 1.21, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["brita water filter", 18100, 0.31, 1, "Tap Filter", "Stable", "", "Tapp vs Brita"],
    ["tap water filter", 12100, 0.86, 1, "Tap Filter", "Stable", "#9", "EcoPro Product Page"],
    ["water filter jug", 12100, 0.58, 1, "Tap Filter", "Stable", "#8", "Best Water Filter Australia 2026"],
    ["faucet water filter", 8100, 0.85, 1, "Tap Filter", "Rising", "", "EcoPro Product Page"],
    ["water filter tap", 8100, 0.94, 1, "Tap Filter", "Stable", "#3", "EcoPro Product Page"],
    ["water purifier", 5400, 0.79, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["under sink water filter", 5400, 0.81, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["whole house water filter", 5400, 1.77, 1, "Tap Filter", "Rising", "", ""],
    ["water filter bottle", 4400, 0.44, 1, "Tap Filter", "Stable", "", ""],
    ["stefani water filter", 4400, 0.06, 1, "Tap Filter", "Stable", "", ""],
    ["water filter australia", 3600, 1.26, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["best water filter australia", 2400, 0.81, 1, "Tap Filter", "Declining", "", "Best Water Filter Australia 2026"],
    ["benchtop water filter", 1900, 0.69, 1, "Tap Filter", "Declining", "", "Best Water Filter Australia 2026"],
    ["water filter cartridge", 1600, 0.47, 1, "Tap Filter", "Stable", "", ""],
    ["alkaline water filter", 1300, 0.97, 1, "Tap Filter", "Declining", "", ""],
    ["on tap water filter", 880, 1.13, 1, "Tap Filter", "Stable", "", "EcoPro Product Page"],
    ["water filter for tap australia", 590, 1.14, 1, "Tap Filter", "Rising", "", "EcoPro Product Page"],
    ["waterdrop filter", 590, 0.31, 1, "Tap Filter", "Rising", "", ""],
    ["best faucet water filter", 480, 1.16, 1, "Tap Filter", "Stable", "#2", "Best Water Filter Australia 2026"],
    ["countertop water filter", 480, 0.92, 1, "Tap Filter", "Rising", "", "Countertop RO Collection"],
    ["best under sink water filter australia", 480, 0.63, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["best water filter jug australia", 390, 0.32, 1, "Tap Filter", "Declining", "", "Best Water Filter Australia 2026"],
    ["doulton water filter", 390, 0.40, 1, "Tap Filter", "Stable", "", ""],
    ["brita replacement filter", 390, 0.28, 1, "Competitor", "Stable", "", "Tapp vs Brita"],
    ["best benchtop water filter australia", 320, 0.68, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["best tap water filter australia", 210, 1.13, 1, "Tap Filter", "Rising", "", "EcoPro Product Page"],
    ["brita on tap", 210, 0.24, 1, "Competitor", "Stable", "", "Tapp vs Brita"],
    ["filtered water bottle australia", 210, 0.55, 1, "Tap Filter", "Declining", "", ""],
    ["screw on tap filter", 140, 0.88, 1, "Tap Filter", "Stable", "#3", "EcoPro Product Page"],
    ["benchtop water purifier", 140, 0.74, 1, "Tap Filter", "Stable", "", "Best Water Filter Australia 2026"],
    ["pur water filter", 140, 0.62, 0.98, "Competitor", "Stable", "", ""],
    ["brita filter review", 110, 0.00, 0.89, "Competitor", "Stable", "", "Tapp vs Brita"],
    ["faucet mount water filter", 90, 1.03, 1, "Tap Filter", "Rising", "", "EcoPro Product Page"],
    ["plastic free water filter", 90, 0.59, 1, "Sustainability", "Stable", "", "Bottled Water vs Tap"],
    ["brita alternative", 10, 0.00, 1, "Competitor", "Stable", "", "Tapp vs Brita"],

    # Shower Filter
    ["shower filter", 12100, 1.13, 1, "Shower Filter", "Stable", "", "Shower Filter Collection"],
    ["shower head filter", 9900, 1.05, 1, "Shower Filter", "Declining", "", "Shower Filter Collection"],
    ["filtered shower head", 9900, 1.05, 1, "Shower Filter", "Declining", "", "Shower Filter Collection"],
    ["shower head water filter", 6600, 0.96, 1, "Shower Filter", "Stable", "", "Shower Filter Collection"],
    ["shower water filter", 2900, 1.13, 1, "Shower Filter", "Stable", "", "Shower Filter Collection"],
    ["shower filter australia", 1300, 1.13, 1, "Shower Filter", "Stable", "", "Shower Filter Collection"],
    ["shower water filter australia", 1300, 0.93, 1, "Shower Filter", "Rising", "", "Shower Filter Collection"],
    ["shower filter for hard water", 1000, 0.88, 1, "Shower Filter", "Rising", "", "Shower Filter for Hard Water"],
    ["best shower filter australia", 480, 0.94, 1, "Shower Filter", "Rising", "", "Best Shower Filter Australia"],
    ["shower head filter australia", 480, 1.22, 1, "Shower Filter", "Declining", "", "Shower Filter Collection"],
    ["best shower filter", 390, 1.10, 0.91, "Shower Filter", "Stable", "", "Best Shower Filter Australia"],
    ["shower filter for chlorine", 390, 0.84, 1, "Shower Filter", "Stable", "", "Shower Filter for Chlorine"],
    ["chlorine shower filter", 390, 0.84, 1, "Shower Filter", "Stable", "", "Shower Filter for Chlorine"],
    ["vitamin c shower filter", 210, 1.04, 1, "Shower Filter", "Stable", "", "Vitamin C Shower Filter"],
    ["inline shower filter", 210, 1.20, 1, "Shower Filter", "Volatile", "", "Shower Filter Collection"],
    ["shower filter for eczema", 70, 0.97, 0.84, "Health/Beauty", "Spiking", "", "Shower Filter Eczema"],
    ["kdf shower filter", 70, 0.99, 1, "Shower Filter", "Volatile", "", "Best Shower Filter Australia"],
    ["shower filter for hair loss", 30, 1.07, 1, "Health/Beauty", "Spiking", "", "Shower Filter Hair Loss"],
    ["shower filter benefits", 30, 0.68, 1, "Shower Filter", "Stable", "", "Shower Filter Collection"],
    ["shower filter for dry skin", 10, 0.00, 1, "Health/Beauty", "Spiking", "", "Shower Filter Eczema"],

    # Reverse Osmosis
    ["reverse osmosis water filter", 8100, 0.81, 1, "Reverse Osmosis", "Rising", "", "Best RO Water Filter Australia"],
    ["reverse osmosis water", 2400, 0.90, 1, "Reverse Osmosis", "Stable", "", "Best RO Water Filter Australia"],
    ["reverse osmosis drinking water", 2400, 0.90, 1, "Reverse Osmosis", "Stable", "", "Best RO Water Filter Australia"],
    ["reverse osmosis water filter australia", 1300, 0.92, 1, "Reverse Osmosis", "Stable", "", "Best RO Water Filter Australia"],
    ["ro water filter", 1000, 0.86, 1, "Reverse Osmosis", "Stable", "", "Best RO Water Filter Australia"],
    ["reverse osmosis system", 880, 1.04, 0.99, "Reverse Osmosis", "Stable", "", "Best RO Water Filter Australia"],
    ["reverse osmosis water filter bunnings", 590, 0.78, 1, "Bunnings", "Declining", "", "Water Filter Bunnings"],
    ["reverse osmosis water filter system", 480, 1.13, 1, "Reverse Osmosis", "Rising", "", "Countertop RO Collection"],
    ["reverse osmosis water purifier", 480, 0.85, 1, "Reverse Osmosis", "Spiking", "", "Countertop RO Collection"],
    ["best reverse osmosis water filter australia", 320, 0.88, 1, "Reverse Osmosis", "Declining", "", "Best RO Water Filter Australia"],
    ["under sink reverse osmosis", 260, 1.02, 1, "Reverse Osmosis", "Stable", "", "Countertop vs Under-Sink RO"],
    ["countertop reverse osmosis", 210, 0.82, 1, "Reverse Osmosis", "Declining", "", "Countertop RO Collection"],
    ["countertop reverse osmosis water filter", 210, 0.95, 1, "Reverse Osmosis", "Spiking", "", "Countertop RO Collection"],
    ["reverse osmosis water filter perth", 170, 1.39, 1, "City - Perth", "Stable", "", "Tap Water Perth"],
    ["reverse osmosis system australia", 140, 1.22, 1, "Reverse Osmosis", "Declining", "", "Best RO Water Filter Australia"],
    ["ro system australia", 140, 1.14, 0.77, "Reverse Osmosis", "Declining", "", "Best RO Water Filter Australia"],
    ["reverse osmosis", 110, 0.93, 1, "Reverse Osmosis", "Rising", "", "Best RO Water Filter Australia"],
    ["reverse osmosis water filter adelaide", 110, 0.94, 1, "Reverse Osmosis", "Volatile", "", ""],
    ["reverse osmosis water filter replacement", 90, 0.91, 1, "Reverse Osmosis", "Rising", "", ""],
    ["benchtop reverse osmosis water filter australia", 90, 0.82, 1, "Reverse Osmosis", "Stable", "", "Countertop RO Collection"],
    ["waterdrop reverse osmosis", 90, 0.51, 1, "Competitor", "Rising", "", "Best RO Water Filter Australia"],
    ["benchtop reverse osmosis water filter", 70, 0.84, 1, "Reverse Osmosis", "Stable", "", "Countertop RO Collection"],
    ["reverse osmosis remineralization", 70, 1.09, 1, "Reverse Osmosis", "Stable", "", "Best RO Water Filter Australia"],
    ["reverse osmosis water filter melbourne", 70, 1.51, 1, "City - Melbourne", "Stable", "", "Tap Water Melbourne"],
    ["tankless reverse osmosis", 70, 1.36, 1, "Reverse Osmosis", "Volatile", "", "Countertop RO Collection"],
    ["countertop water purifier", 50, 0.68, 0.99, "Reverse Osmosis", "Spiking", "", "Countertop RO Collection"],
    ["portable reverse osmosis", 40, 0.73, 1, "Reverse Osmosis", "Stable", "", "Countertop RO Collection"],
    ["best countertop reverse osmosis", 30, 0.50, 1, "Reverse Osmosis", "Stable", "", "Countertop RO Collection"],
    ["countertop ro water filter", 30, 0.77, 1, "Reverse Osmosis", "Stable", "", "Countertop RO Collection"],
    ["countertop ro system", 20, 0.85, 1, "Reverse Osmosis", "Stable", "", "Countertop RO Collection"],
    ["waterdrop k19", 20, 0.55, 1, "Competitor", "Stable", "", "Best RO Water Filter Australia"],
    ["is reverse osmosis water healthy", 20, 0.00, 0.13, "Reverse Osmosis", "Stable", "", "RO vs Tap Filter"],
    ["reverse osmosis vs water filter", 20, 0.55, 0.93, "Reverse Osmosis", "Stable", "", "RO vs Tap Filter"],
    ["reverse osmosis cost", 10, 0.99, 0.85, "Reverse Osmosis", "Stable", "", "Countertop vs Under-Sink RO"],
    ["waterdrop countertop", 10, 0.85, 1, "Competitor", "Rising", "", "Best RO Water Filter Australia"],

    # Contaminants - Fluoride
    ["fluoride in water", 2400, 0.16, 0.06, "Contaminant - Fluoride", "Declining", "", "Fluoride in Water Australia"],
    ["fluoride water filter", 590, 0.99, 1, "Contaminant - Fluoride", "Declining", "", "Fluoride in Water Australia"],
    ["fluoride in water australia", 590, 0.38, 0.04, "Contaminant - Fluoride", "Declining", "", "Fluoride in Water Australia"],
    ["water filter that removes fluoride", 480, 0.92, 1, "Contaminant - Fluoride", "Declining", "", "Fluoride in Water Australia"],
    ["fluoride filter", 390, 0.92, 1, "Contaminant - Fluoride", "Declining", "", "Fluoride in Water Australia"],
    ["remove fluoride from water", 110, 0.86, 1, "Contaminant - Fluoride", "Declining", "", "Fluoride in Water Australia"],

    # Contaminants - Hard Water
    ["limescale", 2400, 0.00, 0.06, "Contaminant - Hard Water", "Declining", "", "Hard Water Australia"],
    ["water softener", 2400, 0.90, 1, "Contaminant - Hard Water", "Declining", "", "Hard Water Australia"],
    ["hard water", 1900, 0.05, 0.04, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],
    ["hard water filter", 590, 0.98, 1, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],
    ["hard water australia", 210, 0.00, 0.07, "Contaminant - Hard Water", "Declining", "", "Hard Water Australia"],
    ["hard water perth", 210, 1.28, 0.02, "Contaminant - Hard Water", "Stable", "", "Tap Water Perth"],
    ["hard water melbourne", 210, 0.77, 0.02, "Contaminant - Hard Water", "Stable", "", "Tap Water Melbourne"],
    ["water hardness australia", 210, 0.00, 0.07, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],
    ["hard water sydney", 140, 0.00, 0.01, "Contaminant - Hard Water", "Stable", "", "Tap Water Sydney"],
    ["hard water hair", 110, 0.00, 0.55, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],
    ["water softener australia", 110, 0.80, 1, "Contaminant - Hard Water", "Volatile", "", "Hard Water Australia"],
    ["limescale filter", 70, 0.63, 0.98, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],
    ["hard water hair loss", 50, 0.00, 0.17, "Contaminant - Hard Water", "Declining", "", "Shower Filter Hair Loss"],
    ["hard water eczema", 30, 0.00, 0.02, "Contaminant - Hard Water", "Stable", "", "Shower Filter Eczema"],
    ["hard water shower", 40, 1.23, 0.92, "Contaminant - Hard Water", "Stable", "", "Shower Filter Hard Water"],
    ["hard water skin", 20, 0.00, 0.02, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],
    ["hard water dry skin", 10, 0.00, 0, "Contaminant - Hard Water", "Stable", "", "Shower Filter Eczema"],
    ["calcium in water", 90, 0.00, 0.11, "Contaminant - Hard Water", "Stable", "", "Hard Water Australia"],

    # Contaminants - PFAS
    ["pfas water filter", 480, 0.89, 1, "Contaminant - PFAS", "Volatile", "#5", "PFAS Water Australia"],
    ["pfas in water", 390, 0.14, 0.18, "Contaminant - PFAS", "Volatile", "", "PFAS Water Australia"],
    ["pfas australia", 320, 0.00, 0.02, "Contaminant - PFAS", "Volatile", "", "PFAS Water Australia"],
    ["pfas filter", 260, 1.43, 1, "Contaminant - PFAS", "Volatile", "", "PFAS Water Australia"],
    ["forever chemicals australia", 90, 0.06, 0.11, "Contaminant - PFAS", "Declining", "", "PFAS Water Australia"],
    ["pfas in drinking water", 70, 1.38, 0.17, "Contaminant - PFAS", "Declining", "", "PFAS Water Australia"],
    ["remove pfas from water", 30, 0.93, 1, "Contaminant - PFAS", "Declining", "", "PFAS Water Australia"],
    ["pfas water treatment", 20, 0.94, 0.57, "Contaminant - PFAS", "Declining", "", "PFAS Water Australia"],
    ["forever chemicals water", 30, 0.00, 0.02, "Contaminant - PFAS", "Stable", "", "PFAS Water Australia"],

    # Contaminants - Microplastics
    ["microplastics water filter", 320, 0.87, 1, "Contaminant - Microplastics", "Rising", "", "Microplastics Water Filter"],
    ["microplastics filter", 140, 0.68, 1, "Contaminant - Microplastics", "Volatile", "", "Microplastics Water Filter"],
    ["microplastics in water", 110, 0.00, 0, "Contaminant - Microplastics", "Stable", "", "Microplastics Water Filter"],
    ["microplastics in bottled water", 90, 0.02, 0.05, "Contaminant - Microplastics", "Stable", "", "Bottled Water vs Tap"],
    ["microplastics in drinking water", 30, 0.00, 0.01, "Contaminant - Microplastics", "Stable", "", "Microplastics Water Filter"],
    ["microplastics in tap water", 20, 0.00, 0.1, "Contaminant - Microplastics", "Stable", "", "Microplastics Water Filter"],
    ["remove microplastics from water", 20, 0.09, 0.58, "Contaminant - Microplastics", "Volatile", "", "Microplastics Water Filter"],

    # Contaminants - Chlorine
    ["chlorine in water", 260, 0.00, 0.11, "Contaminant - Chlorine", "Stable", "", "Chlorine in Tap Water"],
    ["chlorine in tap water", 90, 0.00, 0.04, "Contaminant - Chlorine", "Declining", "", "Chlorine in Tap Water"],
    ["chlorine water filter", 90, 0.96, 1, "Contaminant - Chlorine", "Declining", "", "Chlorine in Tap Water"],
    ["chlorine hair damage", 90, 0.00, 0.37, "Health/Beauty", "Stable", "", "Shower Filter Chlorine"],
    ["chlorine skin", 50, 0.00, 0.04, "Health/Beauty", "Rising", "", "Shower Filter Chlorine"],
    ["remove chlorine from water", 40, 0.69, 1, "Contaminant - Chlorine", "Stable", "", "Chlorine in Tap Water"],

    # Contaminants - Lead
    ["lead in water", 90, 0.00, 0, "Contaminant - Lead", "Stable", "", "Lead in Water Australia"],
    ["lead water filter", 70, 1.17, 0.89, "Contaminant - Lead", "Rising", "", "Lead in Water Australia"],

    # Contaminants - Other
    ["heavy metals in water", 50, 0.00, 0, "Contaminant - Other", "Stable", "", ""],
    ["heavy metals water filter", 50, 1.06, 1, "Contaminant - Other", "Declining", "", ""],
    ["mercury in water", 70, 0.00, 0, "Contaminant - Other", "Stable", "", ""],
    ["nitrates in water", 70, 0.00, 0.01, "Contaminant - Other", "Stable", "", ""],
    ["e coli water", 170, 0.00, 0.01, "Contaminant - Other", "Volatile", "", ""],
    ["bacteria in water", 110, 0.00, 0.01, "Contaminant - Other", "Stable", "", ""],

    # City clusters
    ["tap water melbourne", 590, 0.00, 0.01, "City - Melbourne", "Volatile", "", "Tap Water Melbourne"],
    ["melbourne water quality", 480, 0.00, 0, "City - Melbourne", "Volatile", "", "Tap Water Melbourne"],
    ["can you drink tap water in melbourne", 320, 0.00, 0, "City - Melbourne", "Stable", "", "Tap Water Melbourne"],
    ["is melbourne water safe to drink", 140, 0.00, 0.01, "City - Melbourne", "Stable", "", "Tap Water Melbourne"],
    ["can you drink tap water in australia", 590, 0.00, 0, "City - National", "Stable", "#4", "Can You Drink Tap Water Australia"],
    ["can you drink tap water in sydney", 480, 0.00, 0, "City - Sydney", "Stable", "", "Tap Water Sydney"],
    ["tap water sydney", 390, 0.00, 0.01, "City - Sydney", "Stable", "", "Tap Water Sydney"],
    ["sydney water quality", 390, 0.00, 0, "City - Sydney", "Volatile", "", "Tap Water Sydney"],
    ["is sydney water safe to drink", 260, 0.00, 0, "City - Sydney", "Stable", "", "Tap Water Sydney"],
    ["tap water brisbane", 260, 0.00, 0.01, "City - Brisbane", "Volatile", "", "Tap Water Brisbane"],
    ["brisbane water quality", 260, 0.00, 0.02, "City - Brisbane", "Volatile", "", "Tap Water Brisbane"],
    ["can you drink tap water in brisbane", 170, 0.00, 0, "City - Brisbane", "Stable", "", "Tap Water Brisbane"],
    ["is tap water safe to drink", 170, 0.13, 0.02, "City - National", "Stable", "", "Can You Drink Tap Water Australia"],
    ["tap water perth", 140, 0.00, 0.01, "City - Perth", "Stable", "", "Tap Water Perth"],
    ["is tap water safe to drink in brisbane", 110, 0.00, 0, "City - Brisbane", "Stable", "", "Tap Water Brisbane"],
    ["canberra water quality", 110, 2.88, 0.07, "City - Canberra", "Rising", "", "Tap Water Canberra"],
    ["perth water quality", 110, 0.00, 0.02, "City - Perth", "Stable", "", "Tap Water Perth"],
    ["is perth water safe to drink", 90, 0.00, 0.02, "City - Perth", "Stable", "", "Tap Water Perth"],
    ["is brisbane water safe to drink", 90, 0.00, 0, "City - Brisbane", "Declining", "", "Tap Water Brisbane"],
    ["adelaide water quality", 40, 2.42, 0.25, "City - Adelaide", "Stable", "", ""],

    # Travel
    ["can you drink the tap water in singapore", 260, 0.00, 0.01, "Travel", "Stable", "", ""],
    ["can you drink the tap water in thailand", 90, 0.00, 0, "Travel", "Stable", "", ""],
    ["can you drink the tap water in italy", 90, 0.00, 0.01, "Travel", "Stable", "", ""],
    ["can you drink the tap water in japan", 50, 0.00, 0, "Travel", "Stable", "", ""],

    # Use Case
    ["water filter bunnings", 12100, 0.06, 1, "Bunnings", "Stable", "", "Water Filter Bunnings"],
    ["water filter caravan", 2900, 0.29, 1, "Use Case", "Stable", "", "Water Filter Caravan"],
    ["water test kit", 1600, 0.70, 1, "Use Case", "Stable", "", ""],
    ["water filter fridge", 1300, 0.57, 1, "Use Case", "Stable", "", ""],
    ["water filter camping", 880, 0.16, 1, "Use Case", "Stable", "", "Water Filter Caravan"],
    ["water purification", 590, 1.22, 0.77, "Use Case", "Stable", "", ""],
    ["water quality test", 480, 0.83, 1, "Use Case", "Stable", "", ""],
    ["water filter office", 210, 1.86, 1, "Use Case", "Stable", "", "Water Filter Office"],
    ["water quality australia", 170, 0.00, 0.01, "Use Case", "Declining", "", "Can You Drink Tap Water Australia"],
    ["water treatment australia", 170, 1.45, 0.3, "Use Case", "Declining", "", ""],
    ["how to change water filter", 90, 1.47, 0.06, "Use Case", "Stable", "", ""],
    ["water filter comparison", 70, 0.48, 1, "Use Case", "Declining", "", "RO vs Tap Filter"],
    ["how long do water filters last", 50, 0.78, 0.44, "Use Case", "Stable", "", ""],
    ["water filter subscription", 10, 1.09, 0.71, "Use Case", "Stable", "", ""],
    ["water filter for renters", 10, 1.68, 1, "Use Case", "Stable", "", "Water Filter for Renters"],

    # Sustainability
    ["bottled water plastic", 170, 0.00, 0.71, "Sustainability", "Stable", "", "Bottled Water vs Tap"],
    ["bottled water australia", 140, 0.66, 1, "Sustainability", "Volatile", "", "Bottled Water vs Tap"],
    ["filtered water benefits", 110, 0.05, 0.25, "Sustainability", "Declining", "", "Bottled Water vs Tap"],
    ["is bottled water better than tap", 90, 0.00, 0, "Sustainability", "Stable", "", "Bottled Water vs Tap"],
    ["bpa water bottle", 90, 0.00, 0.76, "Sustainability", "Stable", "", "Bottled Water vs Tap"],
    ["tap water vs filtered water", 70, 0.03, 0.15, "Sustainability", "Volatile", "", "Bottled Water vs Tap"],
    ["what is in tap water", 70, 0.00, 0.02, "Sustainability", "Stable", "", "Can You Drink Tap Water Australia"],
    ["bottled water vs filtered water", 20, 0.00, 0.09, "Sustainability", "Stable", "", "Bottled Water vs Tap"],
    ["water filter vs bottled water", 20, 0.00, 0.01, "Sustainability", "Stable", "", "Bottled Water vs Tap"],

    # Health/Beauty
    ["shower filter for eczema", 70, 0.97, 0.84, "Health/Beauty", "Spiking", "", "Shower Filter Eczema"],
    ["shower filter for hair loss", 30, 1.07, 1, "Health/Beauty", "Spiking", "", "Shower Filter Hair Loss"],
    ["water filter for eczema", 20, 0.85, 1, "Health/Beauty", "Stable", "", "Shower Filter Eczema"],
    ["water filter for hair", 20, 0.89, 0.95, "Health/Beauty", "Stable", "", "Shower Filter Chlorine"],
    ["water filter for skin", 10, 0.00, 0.44, "Health/Beauty", "Stable", "", "Shower Filter Collection"],
    ["drinking water while pregnant", 20, 0.00, 0, "Health/Beauty", "Stable", "", ""],
    ["water filter for baby formula", 10, 0.00, 0.67, "Health/Beauty", "Stable", "", ""],
    ["filtered water for babies", 10, 0.00, 1, "Health/Beauty", "Stable", "", ""],

    # Tapp Branded
    ["tapp water", 150, 0.00, 0, "Branded", "Stable", "#1", "Homepage"],
    ["tapp water filter", 63, 0.00, 0, "Branded", "Stable", "#2", "EcoPro Product Page"],
    ["tapp water filter review", 43, 0.50, 1, "Branded", "Stable", "#1", ""],
    ["tapp water australia", 27, 0.00, 0, "Branded", "Stable", "#1", "Homepage"],
]

for row in keywords:
    ws2.append(row)

style_header(ws2, len(headers2))
style_body(ws2, len(keywords) + 1, len(headers2))

# Apply trend color fills
for row in range(2, len(keywords) + 2):
    trend = ws2.cell(row=row, column=6).value
    if trend in trend_fills:
        fill = trend_fills[trend]
        ws2.cell(row=row, column=6).fill = fill

# Format number columns
for row in range(2, len(keywords) + 2):
    ws2.cell(row=row, column=2).number_format = NUMBER_FMT
    ws2.cell(row=row, column=3).number_format = CURRENCY_FMT

widths2 = [42, 18, 12, 14, 24, 16, 18, 40]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════
# SHEET 3: Competitor Landscape
# ════════════════════════════════════════════
ws3 = wb.create_sheet("Competitor Landscape")

headers3 = ["Competitor", "Domain", "Category Focus", "Est. AU Organic Traffic/mo",
            "Top Keywords", "Strengths", "Weaknesses"]
ws3.append(headers3)

competitors = [
    ["Brita", "brita.com.au", "Water filter jugs, tap filters", 14000,
     "water purification jug (#1, 4,500)\nbrita water filter (#2, 2,700)\nbrita water filter jug (#1, 1,500)\nwater filtration water bottle (#1, 1,300)\nlimescale (#2, 390)",
     "Massive brand awareness\n#1 for jug terms\nStrong retail distribution (Coles, Woolworths, Bunnings)",
     "Weak on tap filter terms\nNo shower filter product\nNo RO product\nNo content marketing"],

    ["Filtered Beauty", "filteredbeauty.com.au", "Shower filters", 4000,
     "shower filter (#2, 1,500)\nwater filtration for shower (#2, 1,500)\nbeauty filter (#1, 430)\nshower water filter (#2, 200)\nshower filter for hair (#1, 120)",
     "Owns shower filter category in AU\nBeauty/wellness positioning\nStrong brand story (hair stylist founder)",
     "Shower-only, no tap or RO\nNarrow product range\nSmall team"],

    ["Waterdrop", "waterdropfilter.com.au", "RO / Under-sink filters", 5500,
     "water drop (#2, 1,800)\nreverse osmosis water filter (#3, 1,200)\nwater filter (#7, 1,000)\nreverse osmosis filter (#3, 880)\ncountertop reverse osmosis (#1, 52)",
     "Owns RO category in AU\nGood blog content driving traffic\nGrowing brand (waterdrop filter 320→880)",
     "US-centric content adapted for AU\nNo shower products\nBrand still building AU awareness"],

    ["My Water Filter", "mywaterfilter.com.au", "All categories (retailer)", 7000,
     "water filter (#3, 3,000)\nwater filters (#1, 1,200)\nwater filtration for shower (#4, 600)\nwater filtration system (#6, 410)\nshower filter (#6, 310)",
     "Broadest authority across all categories\n#1 for benchtop RO AU\nRetailer model with wide range",
     "Retailer, not a brand\nDated website\nNo single brand focus\nWeak content quality"],

    ["Water Filters Australia", "waterfiltersaustralia.com.au", "RO / Countertop", 800,
     "water filters australia (#3, 430)\nreverse osmosis filter countertop (#1, 64)\ncountertop RO water filter (#2, 34)\nreverse osmosis benchtop (#1, 21)",
     "Owns countertop/benchtop RO niche terms\nAU-focused domain name",
     "Small traffic overall\nLimited product range\nLow domain authority"],

    ["Well Verti", "wellverti.com.au", "Shower filters", 300,
     "shower water filter (#6, 41)\nshower head filter hard water AU (#2, 23)\nbest water filter for shower heads (#3, 20)\nhard water shower filter (#2, 14)",
     "Claims 80K+ customers\nStrong social/influencer marketing\nProductReview.com.au presence",
     "Very weak organic presence\nMost traffic from paid/social\nNot building SEO moat"],

    ["Fettle Effect", "fettleeffect.com.au", "Shower filters", 100,
     "fettle (#4, 66)\nbest shower head filter australia (#2, 18)",
     "Beauty/wellness angle\nSalon partnerships\n15-stage filter marketing",
     "Tiny organic footprint\nNarrow keyword coverage\nHeavily reliant on offline/social"],

    ["Stefani", "N/A", "Benchtop ceramic filters", 0,
     "stefani water filter (4,400 branded search)",
     "Massive branded search volume\nHeritage Australian brand\nRetail distribution",
     "No meaningful non-branded organic\nDated brand perception\nNo digital content strategy"],

    ["Watego", "watego.com.au", "Shower filters", 0,
     "shower filter for hair australia (#12, 0)",
     "Active blog content strategy\nComparison posts (Watego vs Well Verti)",
     "Essentially invisible organically\nContent not ranking\nNew entrant"],

    ["Tapp Water (current)", "tappwater.com.au", "Tap filters, shower filters, RO (new)", 1500,
     "tap filter (#4, 180)\ntapp water (#1, 150)\nbest tap filter (#1, 120)\ntapp water filter (#2, 63)\ntapp water filter review (#1, 43)",
     "Strong branded rankings\n#1 for 'best tap filter'\nThree-product lineup (tap + shower + RO)\nSubscription model = LTV",
     "Low overall organic traffic\nWeak shower filter rankings (#8 only)\nNo city/contaminant content\nNew to RO category"],
]

for row in competitors:
    ws3.append(row)

style_header(ws3, len(headers3))
style_body(ws3, len(competitors) + 1, len(headers3))

for row in range(2, len(competitors) + 2):
    ws3.cell(row=row, column=4).number_format = NUMBER_FMT

# Highlight Tapp row
tapp_row = len(competitors) + 1
for col in range(1, len(headers3) + 1):
    ws3.cell(row=tapp_row, column=col).fill = PatternFill('solid', fgColor='E8F0FE')
    ws3.cell(row=tapp_row, column=col).font = BOLD_FONT

widths3 = [18, 30, 28, 22, 50, 50, 50]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════════
# SHEET 4: Projections
# ════════════════════════════════════════════
ws4 = wb.create_sheet("Projections")

headers4 = ["Page Title", "Phase", "Cluster Volume",
            "Conservative (Pos 5-10, 3% CTR)", "Moderate (Pos 3-5, 8% CTR)",
            "Aggressive (Pos 1-3, 15% CTR)",
            "Rev Conservative ($)", "Rev Moderate ($)", "Rev Aggressive ($)"]
ws4.append(headers4)

# page data: title, phase, cluster vol
proj_data = [
    ("Shower Filter Collection", 1, 44000),
    ("Countertop RO Collection", 1, 670),
    ("EcoPro Product Page", 1, 29180),
    ("Best Water Filter Australia 2026", 1, 6210),
    ("Best Shower Filter Australia 2026", 1, 2170),
    ("Best RO Water Filter Australia 2026", 1, 9720),
    ("Tap Water Sydney", 2, 1660),
    ("Tap Water Melbourne", 2, 1740),
    ("Tap Water Brisbane", 2, 890),
    ("Tap Water Perth", 2, 720),
    ("Can You Drink Tap Water in Australia", 2, 880),
    ("Tap Water Canberra", 2, 110),
    ("PFAS Water Australia", 2, 1640),
    ("Fluoride in Water Australia", 2, 4560),
    ("Microplastics Water Filter", 2, 730),
    ("Hard Water Australia", 2, 8414),
    ("Chlorine in Tap Water", 2, 620),
    ("Shower Filter Hard Water", 3, 1040),
    ("Shower Filter Chlorine", 3, 920),
    ("Shower Filter Eczema", 3, 120),
    ("Shower Filter Hair Loss", 3, 190),
    ("RO vs Tap Filter", 3, 160),
    ("Countertop vs Under-Sink RO", 3, 470),
    ("Water Filter for Renters", 3, 50),
    ("Water Filter Caravan", 3, 3780),
    ("Tapp vs Brita", 3, 18220),
    ("Water Filter Bunnings", 3, 12690),
    ("Bottled Water vs Tap", 3, 530),
    ("Water Filter Office", 3, 210),
    ("Lead in Water Australia", 3, 160),
    ("Vitamin C Shower Filter", 3, 210),
]

for i, (title, phase, vol) in enumerate(proj_data, 2):
    ws4.cell(row=i, column=1, value=title)
    ws4.cell(row=i, column=2, value=phase)
    ws4.cell(row=i, column=3, value=vol)
    # Conservative: 3% CTR
    ws4.cell(row=i, column=4).value = f'=C{i}*0.03'
    # Moderate: 8% CTR
    ws4.cell(row=i, column=5).value = f'=C{i}*0.08'
    # Aggressive: 15% CTR
    ws4.cell(row=i, column=6).value = f'=C{i}*0.15'
    # Revenue = clicks * 3% conversion * $100 AOV
    ws4.cell(row=i, column=7).value = f'=D{i}*0.03*100'
    ws4.cell(row=i, column=8).value = f'=E{i}*0.03*100'
    ws4.cell(row=i, column=9).value = f'=F{i}*0.03*100'

last_row = len(proj_data) + 1
total_row = last_row + 1

# Phase subtotals
ws4.cell(row=total_row, column=1, value="PHASE 1 TOTAL")
ws4.cell(row=total_row, column=1).font = BOLD_FONT
for col in range(3, 10):
    ws4.cell(row=total_row, column=col).value = f'=SUMPRODUCT((B2:B{last_row}=1)*{get_column_letter(col)}2:{get_column_letter(col)}{last_row})'
    ws4.cell(row=total_row, column=col).font = BOLD_FONT

total_row += 1
ws4.cell(row=total_row, column=1, value="PHASE 2 TOTAL")
ws4.cell(row=total_row, column=1).font = BOLD_FONT
for col in range(3, 10):
    ws4.cell(row=total_row, column=col).value = f'=SUMPRODUCT((B2:B{last_row}=2)*{get_column_letter(col)}2:{get_column_letter(col)}{last_row})'
    ws4.cell(row=total_row, column=col).font = BOLD_FONT

total_row += 1
ws4.cell(row=total_row, column=1, value="PHASE 3 TOTAL")
ws4.cell(row=total_row, column=1).font = BOLD_FONT
for col in range(3, 10):
    ws4.cell(row=total_row, column=col).value = f'=SUMPRODUCT((B2:B{last_row}=3)*{get_column_letter(col)}2:{get_column_letter(col)}{last_row})'
    ws4.cell(row=total_row, column=col).font = BOLD_FONT

total_row += 2
ws4.cell(row=total_row, column=1, value="GRAND TOTAL")
ws4.cell(row=total_row, column=1).font = Font(name='Arial', bold=True, size=12)
for col in range(3, 10):
    ws4.cell(row=total_row, column=col).value = f'=SUM({get_column_letter(col)}{total_row-4}:{get_column_letter(col)}{total_row-2})'
    ws4.cell(row=total_row, column=col).font = Font(name='Arial', bold=True, size=12)

style_header(ws4, len(headers4))
style_body(ws4, last_row, len(headers4))

# Format numbers
for row in range(2, total_row + 1):
    ws4.cell(row=row, column=3).number_format = NUMBER_FMT
    ws4.cell(row=row, column=4).number_format = NUMBER_FMT
    ws4.cell(row=row, column=5).number_format = NUMBER_FMT
    ws4.cell(row=row, column=6).number_format = NUMBER_FMT
    ws4.cell(row=row, column=7).number_format = '$#,##0'
    ws4.cell(row=row, column=8).number_format = '$#,##0'
    ws4.cell(row=row, column=9).number_format = '$#,##0'

# Highlight totals
for r in range(last_row + 1, total_row + 1):
    for col in range(1, 10):
        ws4.cell(row=r, column=col).fill = PatternFill('solid', fgColor='D6E4F0')

# Grand total darker
for col in range(1, 10):
    ws4.cell(row=total_row, column=col).fill = PatternFill('solid', fgColor='B4C7E7')

widths4 = [45, 8, 15, 22, 22, 22, 18, 18, 18]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A2'

# Save
output = '/Users/Shared/tapp-water-au-seo-content-plan.xlsx'
wb.save(output)
print(f"Saved to {output}")
